"""Executive deliverables: PDF report (matplotlib PdfPages) + Excel workbook.

The PDF cover carries the synthetic-data disclaimer and the headline numbers;
inner pages show the detector comparison (PR curves), the health-index ranking
and the before/after maintenance schedule. The Excel workbook has Machines,
Alerts, HealthIndex, Schedule and Comparison sheets.

Chart styling follows a validated light-surface palette: categorical slots
blue #2a78d6 / green #008300 (fixed order), muted ink for axes, hairline grid.
"""

from __future__ import annotations

import math
from pathlib import Path

import matplotlib
import numpy as np

matplotlib.use("Agg")

import matplotlib.pyplot as plt  # noqa: E402
from matplotlib.backends.backend_pdf import PdfPages  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font  # noqa: E402

from pdm.alert_economics import AlertEconomics, curve_rows, economics_from_result  # noqa: E402
from pdm.data import FAULT_HEALTHY  # noqa: E402
from pdm.pipeline import PipelineResult  # noqa: E402
from pdm.rul import RULEvaluation, evaluate_rul  # noqa: E402
from pdm.schedule import ScheduleResult  # noqa: E402

SERIES_3 = "#c25a00"  # amber — the detector's current operating threshold

INK = "#0b0b0b"
INK_2 = "#52514e"
MUTED = "#898781"
GRID = "#e1e0d9"
BASELINE = "#c3c2b7"
SERIES_1 = "#2a78d6"  # blue — PCA / primary
SERIES_2 = "#008300"  # green — autoencoder
SURFACE = "#fcfcfb"

DISCLAIMER = (
    "All data in this report is SYNTHETIC, produced by a seeded generator\n"
    "(pdm/data.py). No real plant telemetry was used. The health index is a\n"
    "heuristic degradation score, not a certified remaining-useful-life\n"
    "prediction. Detector metrics are measured against the generator's\n"
    "ground-truth fault labels."
)


def _style_axis(ax) -> None:
    ax.set_facecolor(SURFACE)
    for side in ("top", "right"):
        ax.spines[side].set_visible(False)
    for side in ("left", "bottom"):
        ax.spines[side].set_color(BASELINE)
    ax.tick_params(colors=MUTED, labelsize=8)
    ax.grid(True, color=GRID, linewidth=0.6, alpha=0.9)
    ax.set_axisbelow(True)


def _fmt(x: float, digits: int = 3) -> str:
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return "n/a"
    return f"{x:.{digits}f}"


def _cover_page(pdf: PdfPages, result: PipelineResult) -> None:
    fig = plt.figure(figsize=(8.27, 11.69))
    fig.patch.set_facecolor(SURFACE)
    rec = result.reports[result.recommended]
    sched = result.schedule
    fig.text(0.08, 0.90, "Predictive Maintenance Report", fontsize=22, color=INK, weight="bold")
    fig.text(
        0.08,
        0.865,
        "Sensor anomaly detection + optimized maintenance scheduling (synthetic fleet)",
        fontsize=11,
        color=INK_2,
    )
    lines = [
        ("Recommended detector", result.recommended.upper()),
        ("  ROC-AUC / PR-AUC", f"{_fmt(rec.roc_auc)} / {_fmt(rec.pr_auc)}"),
        (
            "  Machines detected",
            f"{rec.n_detected}/{rec.n_faulty}"
            f" (mean {_fmt(rec.mean_delay_days, 1)} days after true onset)",
        ),
        (
            "  Threshold",
            f"{rec.fpr_budget:.0%} FPR budget (measured {rec.val_fpr:.1%})",
        ),
        (
            "Schedule vs FIFO baseline",
            f"-{sched.reduction_pct:.1f}% weighted delay"
            f" ({sched.fifo.weighted_delay} -> {sched.optimized.weighted_delay})",
        ),
        ("  Solver status", f"CP-SAT {sched.optimized.status}"),
        (
            "Fleet",
            f"{result.data.config.n_machines} machines x {result.data.n_days} days,"
            f" {len(result.data.faulty_machines())} with injected faults",
        ),
    ]
    y = 0.78
    for label, value in lines:
        fig.text(0.08, y, label, fontsize=11, color=INK_2)
        fig.text(0.40, y, value, fontsize=11, color=INK, weight="bold")
        y -= 0.035
    fig.text(0.08, 0.30, "Data disclaimer", fontsize=12, color=INK, weight="bold")
    fig.text(0.08, 0.22, DISCLAIMER, fontsize=9, color=INK_2, linespacing=1.6)
    pdf.savefig(fig)
    plt.close(fig)


def _detector_page(pdf: PdfPages, result: PipelineResult) -> None:
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(11.69, 5.5))
    fig.patch.set_facecolor(SURFACE)
    colors = {"pca": SERIES_1, "autoencoder": SERIES_2}
    for name, rep in result.reports.items():
        if rep.recall_curve.size:
            ax1.plot(
                rep.recall_curve,
                rep.precision_curve,
                color=colors.get(name, SERIES_1),
                linewidth=2,
                label=f"{name} (PR-AUC {_fmt(rep.pr_auc)})",
            )
    ax1.set_xlabel("Recall", color=INK_2)
    ax1.set_ylabel("Precision", color=INK_2)
    ax1.set_title("Precision-recall on held-out machine-days", color=INK, fontsize=11)
    ax1.set_xlim(0, 1)
    ax1.set_ylim(0, 1.02)
    _style_axis(ax1)
    ax1.legend(frameon=False, fontsize=9, labelcolor=INK_2)

    metric_names = ["ROC-AUC", "PR-AUC", "precision@k"]
    width = 0.35
    for i, (name, rep) in enumerate(result.reports.items()):
        vals = [rep.roc_auc, rep.pr_auc, rep.precision_at_k]
        xs = [j + (i - 0.5) * width for j in range(len(metric_names))]
        bars = ax2.bar(
            xs, vals, width=width * 0.94, color=colors.get(name, SERIES_1), label=name
        )
        for b, v in zip(bars, vals, strict=True):
            ax2.text(
                b.get_x() + b.get_width() / 2,
                v + 0.015,
                _fmt(v),
                ha="center",
                fontsize=8,
                color=INK_2,
            )
    ax2.set_xticks(range(len(metric_names)))
    ax2.set_xticklabels(metric_names, color=INK_2)
    ax2.set_ylim(0, 1.1)
    ax2.set_title(
        f"Detector comparison (recommended: {result.recommended})", color=INK, fontsize=11
    )
    _style_axis(ax2)
    ax2.legend(frameon=False, fontsize=9, labelcolor=INK_2)
    fig.tight_layout()
    pdf.savefig(fig)
    plt.close(fig)


def _health_page(pdf: PdfPages, result: PipelineResult) -> None:
    fig, ax = plt.subplots(figsize=(11.69, 6.5))
    fig.patch.set_facecolor(SURFACE)
    order = result.health.ranking
    truth = {lab.machine: lab.fault_type for lab in result.data.labels}
    names = [f"M{m:02d}" for m in order]
    vals = [result.health.final[m] for m in order]
    ys = range(len(order))
    ax.barh(ys, vals, color=SERIES_1, height=0.62)
    for y, m, v in zip(ys, order, vals, strict=True):
        tag = truth[m]
        note = "" if tag == FAULT_HEALTHY else f"  (true fault: {tag})"
        ax.text(v + 1.0, y, f"{v:.0f}{note}", va="center", fontsize=8, color=INK_2)
    ax.set_yticks(list(ys))
    ax.set_yticklabels(names, color=INK_2, fontsize=8)
    ax.invert_yaxis()
    ax.set_xlim(0, 118)
    ax.set_xlabel("Health index (100 = healthy) - heuristic degradation score, not RUL", color=INK_2)
    ax.set_title("Fleet health ranking (most urgent first)", color=INK, fontsize=11)
    _style_axis(ax)
    fig.tight_layout()
    pdf.savefig(fig)
    plt.close(fig)


def _gantt(ax, sched: ScheduleResult, n_crews: int, n_days: int, day_hours: int, title: str):
    for a in sched.assignments:
        x = a.day * day_hours + a.start_h
        ax.broken_barh(
            [(x, a.duration_h)],
            (a.crew - 0.3, 0.6),
            facecolors=SERIES_1,
            edgecolor=SURFACE,
            linewidth=1.5,
        )
        ax.text(
            x + a.duration_h / 2,
            a.crew,
            f"M{a.machine:02d}",
            ha="center",
            va="center",
            fontsize=8,
            color=SURFACE,
            weight="bold",
        )
    for d in range(1, n_days):
        ax.axvline(d * day_hours, color=BASELINE, linewidth=0.8, linestyle="--")
    ax.set_yticks(range(n_crews))
    ax.set_yticklabels([f"Crew {c + 1}" for c in range(n_crews)], color=INK_2, fontsize=9)
    ax.set_ylim(-0.7, n_crews - 0.3)
    ax.set_xlim(0, n_days * day_hours)
    ax.set_xlabel("Working hours from start of horizon (dashed = day boundary)", color=INK_2)
    ax.set_title(title, color=INK, fontsize=10)
    _style_axis(ax)
    ax.grid(axis="y", visible=False)


def _schedule_page(pdf: PdfPages, result: PipelineResult) -> None:
    sched = result.schedule
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(11.69, 7.0))
    fig.patch.set_facecolor(SURFACE)
    _gantt(
        ax1,
        sched.fifo,
        sched.n_crews,
        sched.n_days,
        sched.day_hours,
        f"FIFO baseline - weighted delay {sched.fifo.weighted_delay}",
    )
    _gantt(
        ax2,
        sched.optimized,
        sched.n_crews,
        sched.n_days,
        sched.day_hours,
        f"CP-SAT optimized ({sched.optimized.status}) - weighted delay "
        f"{sched.optimized.weighted_delay} ({sched.reduction_pct:.1f}% lower)",
    )
    fig.tight_layout()
    pdf.savefig(fig)
    plt.close(fig)


def _alert_economics_page(pdf: PdfPages, econ: AlertEconomics) -> None:
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(11.69, 5.5))
    fig.patch.set_facecolor(SURFACE)
    pts = sorted(econ.points, key=lambda p: p.alerts_per_period)
    load = [p.alerts_per_period for p in pts]

    # left: expected-cost curve vs analyst load, with the two operating points
    ax1.plot([p.alerts_per_period for p in pts], [p.expected_cost for p in pts],
             color=SERIES_1, linewidth=2)
    b = econ.best
    ax1.scatter([b.alerts_per_period], [b.expected_cost], color=SERIES_2, zorder=5, s=45,
                label=f"cost-optimal (thr {b.threshold:.3f})")
    if econ.current is not None:
        c = econ.current
        ax1.scatter([c.alerts_per_period], [c.expected_cost], color=SERIES_3, zorder=5, s=45,
                    label=f"current 5% FPR (thr {c.threshold:.3f})")
    ax1.set_xlabel("Alerts per day (analyst workload -> alert fatigue)", color=INK_2)
    ax1.set_ylabel(f"Expected cost ({econ.cost.unit})", color=INK_2)
    ax1.set_title("Cost vs analyst load (ILLUSTRATIVE costs)", color=INK, fontsize=11)
    _style_axis(ax1)
    ax1.legend(frameon=False, fontsize=8, labelcolor=INK_2)

    # right: the coverage/fatigue trade-off — missed failures and false alarms vs load
    ax2.plot(load, [p.missed_failures for p in pts], color=SERIES_1, linewidth=2,
             label="missed failure-days")
    ax2.plot(load, [p.fp for p in pts], color=SERIES_2, linewidth=2, label="false alarms")
    ax2.axvline(b.alerts_per_period, color=SERIES_2, linewidth=1, linestyle="--", alpha=0.7)
    if econ.current is not None:
        ax2.axvline(econ.current.alerts_per_period, color=SERIES_3, linewidth=1,
                    linestyle="--", alpha=0.7)
    ax2.set_xlabel("Alerts per day", color=INK_2)
    ax2.set_ylabel("Count over the test window", color=INK_2)
    ax2.set_title("Coverage vs fatigue trade-off", color=INK, fontsize=11)
    _style_axis(ax2)
    ax2.legend(frameon=False, fontsize=8, labelcolor=INK_2)

    note = (
        f"SYNTHETIC data; cost rates ILLUSTRATIVE (missed failure = "
        f"{econ.cost.cost_missed_failure:g}, false alarm = {econ.cost.cost_false_alarm:g}, "
        f"ratio {econ.cost.ratio:g}:1) — not a business guarantee. Cost-minimising threshold "
        f"moves with the ratio."
    )
    fig.text(0.08, 0.02, note, fontsize=8, color=MUTED)
    fig.suptitle("Alert-threshold economics / alert-fatigue analysis", fontsize=13,
                 color=INK, weight="bold", x=0.08, ha="left")
    fig.tight_layout(rect=(0, 0.03, 1, 0.95))
    pdf.savefig(fig)
    plt.close(fig)


def _rul_page(pdf: PdfPages, ev: RULEvaluation) -> None:
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(11.69, 5.5))
    fig.patch.set_facecolor(SURFACE)
    c = ev.config
    axis_top = max(float(ev.true_rul.max()), float(ev.pred_rul.max()), 1.0) * 1.05

    # left: predicted vs true RUL with the ideal line and the alpha-lambda cone
    ft = {mr.machine: mr.fault_type for mr in ev.machines}
    is_accel = np.array([ft[int(m)] == "accelerating_drift" for m in ev.sample_machine])
    ax1.plot([0, axis_top], [0, axis_top], color=MUTED, linewidth=1.2, linestyle="--",
             label="ideal")
    grid = np.linspace(0, axis_top, 50)
    ax1.fill_between(grid, np.clip(grid - (c.alpha * grid + c.alpha_floor_days), 0, None),
                     grid + (c.alpha * grid + c.alpha_floor_days), color=SERIES_2, alpha=0.08,
                     label=f"+/-({c.alpha:g}*RUL+{c.alpha_floor_days:g}d) cone")
    ax1.scatter(ev.true_rul[~is_accel], ev.pred_rul[~is_accel], s=18, color=SERIES_1,
                alpha=0.75, label="slow_drift")
    ax1.scatter(ev.true_rul[is_accel], ev.pred_rul[is_accel], s=18, color=SERIES_2,
                alpha=0.75, label="accelerating_drift")
    ax1.set_xlim(0, axis_top)
    ax1.set_ylim(0, axis_top)
    ax1.set_xlabel("True RUL (days to failure)", color=INK_2)
    ax1.set_ylabel("Predicted RUL (days)", color=INK_2)
    ax1.set_title("Predicted vs true RUL (leave-one-machine-out)", color=INK, fontsize=11)
    _style_axis(ax1)
    ax1.legend(frameon=False, fontsize=8, labelcolor=INK_2)

    # right: headline metrics as a small table
    ax2.axis("off")
    rows = [
        ("MAE", f"{ev.mae:.2f} d"),
        ("RMSE", f"{ev.rmse:.2f} d"),
        ("naive-mean baseline MAE", f"{ev.baseline_mae:.2f} d"),
        (f"MAE (RUL <= {c.horizon_days} d)", f"{ev.mae_within_horizon:.2f} d"),
        (f"alpha-accuracy (alpha={c.alpha:g})", f"{ev.alpha_accuracy:.2f}"),
        ("mean prognostic horizon", f"{ev.mean_prognostic_horizon:.1f} d"),
        ("scored machine-days", f"{ev.n_samples}"),
        ("machines (progressive faults)", f"{ev.n_machines}"),
    ]
    y = 0.92
    ax2.text(0.0, y, "Remaining-useful-life accuracy", fontsize=12, color=INK, weight="bold",
             transform=ax2.transAxes)
    y -= 0.10
    for label, value in rows:
        ax2.text(0.0, y, label, fontsize=10, color=INK_2, transform=ax2.transAxes)
        ax2.text(0.72, y, value, fontsize=10, color=INK, weight="bold", transform=ax2.transAxes)
        y -= 0.075
    note = (
        f"SYNTHETIC data. Failure level is ILLUSTRATIVE ({c.vib_rise:g} mm/s of induced "
        f"vibration rise), not a measured engineering limit; day counts scale with it. The "
        f"predictor reuses the recommended detector's own risk score and never sees the true "
        f"onset. Progressive faults (slow_drift, accelerating_drift) only."
    )
    ax2.text(0.0, 0.06, note, fontsize=7.5, color=MUTED, transform=ax2.transAxes, wrap=True)

    fig.suptitle("Remaining-useful-life (RUL) estimation", fontsize=13, color=INK,
                 weight="bold", x=0.08, ha="left")
    fig.tight_layout(rect=(0, 0, 1, 0.95))
    pdf.savefig(fig)
    plt.close(fig)


def export_pdf(result: PipelineResult, path: str | Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    econ = economics_from_result(result)
    rul = evaluate_rul(result)
    with PdfPages(path) as pdf:
        _cover_page(pdf, result)
        _detector_page(pdf, result)
        _health_page(pdf, result)
        _schedule_page(pdf, result)
        _alert_economics_page(pdf, econ)
        _rul_page(pdf, rul)
    return path


def _sheet(wb: Workbook, title: str, header: list[str], rows: list[list]) -> None:
    ws = wb.create_sheet(title)
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"


def export_excel(result: PipelineResult, path: str | Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    truth = {lab.machine: (lab.fault_type, lab.onset_day) for lab in result.data.labels}
    rank_pos = {m: i + 1 for i, m in enumerate(result.health.ranking)}
    _sheet(
        wb,
        "Machines",
        ["machine", "true_fault_type", "true_onset_day", "final_health", "urgency", "rank"],
        [
            [
                m,
                truth[m][0],
                truth[m][1],
                round(float(result.health.final[m]), 2),
                round(float(result.health.urgency[m]), 2),
                rank_pos[m],
            ]
            for m in range(result.data.config.n_machines)
        ],
    )
    _sheet(
        wb,
        "Alerts",
        ["machine", "day", "score", "threshold", "flagged_on"],
        [
            [a.machine, a.day, round(a.score, 5), round(result.chosen_threshold, 5),
             " + ".join(a.sensors)]
            for a in result.alerts
        ],
    )
    _sheet(
        wb,
        "HealthIndex",
        ["machine", "day", "anomaly_score", "health"],
        [
            [m, int(d), round(float(result.chosen_scores[m, i]), 5),
             round(float(result.health.health[m, i]), 2)]
            for m in range(result.data.config.n_machines)
            for i, d in enumerate(result.features.days)
        ],
    )
    weights = {j.machine: j.weight for j in result.schedule.jobs}
    sched_rows = []
    for plan, res in (("FIFO", result.schedule.fifo), ("OPTIMIZED", result.schedule.optimized)):
        for a in res.assignments:
            sched_rows.append(
                [plan, a.machine, a.crew + 1, a.day + 1, a.start_h, a.duration_h,
                 a.completion_h, weights[a.machine]]
            )
    _sheet(
        wb,
        "Schedule",
        ["plan", "machine", "crew", "day", "start_hour", "duration_h",
         "completion_h", "urgency_weight"],
        sched_rows,
    )
    comp_rows = [
        ["data", "synthetic seeded generator (pdm/data.py); no real telemetry", ""],
        ["health index", "heuristic degradation score, NOT certified RUL", ""],
        ["", "", ""],
        ["metric", "pca", "autoencoder"],
    ]
    ae = result.reports.get("autoencoder")
    pca = result.reports["pca"]
    for label, attr in (
        ("roc_auc", "roc_auc"),
        ("pr_auc", "pr_auc"),
        ("precision_at_k", "precision_at_k"),
        ("mean_delay_days", "mean_delay_days"),
        ("machines_detected", "n_detected"),
        ("val_fpr", "val_fpr"),
    ):
        pv = getattr(pca, attr)
        av = getattr(ae, attr) if ae else "n/a"
        comp_rows.append([label, round(pv, 4) if isinstance(pv, float) else pv,
                          round(av, 4) if isinstance(av, float) else av])
    comp_rows += [
        ["recommended", result.recommended,
         f"margin (ae - pca PR-AUC): {'n/a' if ae is None else round(result.margin, 4)}"],
        ["", "", ""],
        ["schedule", "FIFO baseline", "CP-SAT"],
        ["weighted_delay", result.schedule.fifo.weighted_delay,
         result.schedule.optimized.weighted_delay],
        ["reduction_pct", "", round(result.schedule.reduction_pct, 2)],
        ["solver_status", "", result.schedule.optimized.status],
    ]
    _sheet(wb, "Comparison", ["item", "value_a", "value_b"], comp_rows)

    econ = economics_from_result(result)
    econ_rows = [
        ["# SYNTHETIC data; cost rates are ILLUSTRATIVE assumptions, not a business guarantee",
         "", "", "", "", "", "", ""],
        [f"cost_missed_failure={econ.cost.cost_missed_failure:g}",
         f"cost_false_alarm={econ.cost.cost_false_alarm:g}",
         f"ratio={econ.cost.ratio:g}:1", "", "", "", "", ""],
    ]
    for p in curve_rows(econ):
        note = "cost-optimal" if p.threshold == econ.best.threshold else ""
        if econ.current is not None and p.threshold == econ.current.threshold:
            note = (note + "+current").strip("+")
        econ_rows.append([
            round(p.threshold, 5),
            round(p.precision, 4) if p.precision == p.precision else "n/a",
            round(p.recall, 4) if p.recall == p.recall else "n/a",
            p.alerts,
            round(p.alerts_per_period, 3),
            p.missed_failures,
            round(p.expected_cost, 1),
            note,
        ])
    _sheet(
        wb,
        "AlertEconomics",
        ["threshold", "precision", "recall", "alerts", "alerts_per_day",
         "missed_failures", "expected_cost", "note"],
        econ_rows,
    )

    ev = evaluate_rul(result)
    rul_rows = [
        ["# SYNTHETIC data; failure level is ILLUSTRATIVE, not a measured engineering limit",
         "", "", "", "", ""],
        [f"failure_vib_rise={ev.config.vib_rise:g}mm/s",
         f"model=log1p(RUL)=a+b*severity (a={ev.coef[0]:.3f}, b={ev.coef[1]:.3f})",
         "leave-one-machine-out", "", "", ""],
        ["metric", "value", "", "", "", ""],
        ["MAE_days", round(ev.mae, 3), "", "", "", ""],
        ["RMSE_days", round(ev.rmse, 3), "", "", "", ""],
        ["naive_mean_baseline_MAE_days", round(ev.baseline_mae, 3), "", "", "", ""],
        [f"MAE_within_{ev.config.horizon_days}d", round(ev.mae_within_horizon, 3), "", "", "", ""],
        [f"alpha{ev.config.alpha:g}_accuracy", round(ev.alpha_accuracy, 3), "", "", "", ""],
        ["mean_prognostic_horizon_days", round(ev.mean_prognostic_horizon, 2), "", "", "", ""],
        ["scored_machine_days", ev.n_samples, "", "", "", ""],
        ["", "", "", "", "", ""],
        ["per-machine (held-out)", "", "", "", "", ""],
        ["machine", "fault_type", "onset_day", "failure_day", "scored_days", "mae_days"],
    ]
    for mr in ev.machines:
        rul_rows.append(
            [mr.machine, mr.fault_type, mr.onset_day, mr.failure_day, mr.n_days, round(mr.mae, 3)]
        )
    _sheet(
        wb,
        "RUL",
        ["item", "value_b", "value_c", "value_d", "value_e", "value_f"],
        rul_rows,
    )

    wb.save(path)
    return path


def build_deliverables(result: PipelineResult, outdir: str | Path) -> dict[str, int]:
    outdir = Path(outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    pdf = export_pdf(result, outdir / "pdm_report.pdf")
    xlsx = export_excel(result, outdir / "pdm_workbook.xlsx")
    return {str(pdf): pdf.stat().st_size, str(xlsx): xlsx.stat().st_size}
